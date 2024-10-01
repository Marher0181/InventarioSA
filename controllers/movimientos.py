from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, Blueprint
from models import db
from io import BytesIO
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from sqlalchemy import text

movimientos_bp = Blueprint('movimientos', __name__)

@movimientos_bp.route('/movimientos', methods=['GET'])
def gestionar_movimientos():
    if 'usuarioSesion' not in session:
        flash("Debes iniciar sesión para acceder a esta página.")
        return redirect(url_for('auth.login'))
    movimientos = db.session.execute(text("SELECT * FROM MovimientosInventario")).fetchall()
    return render_template('gestionar_movimientos.html', movimientos=movimientos)

@movimientos_bp.route('/reporte-productos-mas-vendidos', methods=['GET', 'POST'])
def reporte_productos_mas_vendidos():
    if request.method == 'POST':
        mes = request.form['mes']
        anio = request.form['anio']
        
        productos = db.session.execute(text("""
            EXEC ObtenerProductosMasVendidosPorMesYAño :mes, :anio
        """), {'mes': mes, 'anio': anio}).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Productos Más Vendidos"
        ws.append(["Producto", "Cantidad Vendida"])

        for producto in productos:
            ws.append([producto[0], producto[1]])

        # Crear gráfico
        chart = BarChart()
        chart.title = "Top Productos Más Vendidos"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(productos) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(productos) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E5")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="reporte_productos_mas_vendidos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template('form_reporte.html', tipo='Productos Más Vendidos')

@movimientos_bp.route('/reporte-productos-menos-vendidos', methods=['GET', 'POST'])
def reporte_productos_menos_vendidos():
    if request.method == 'POST':
        mes = request.form['mes']
        anio = request.form['anio']
        
        productos = db.session.execute(text("""
            EXEC ObtenerProductosMenosVendidosPorMesYAño :mes, :anio
        """), {'mes': mes, 'anio': anio}).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Productos Menos Vendidos"
        ws.append(["Producto", "Cantidad Vendida"])

        for producto in productos:
            ws.append([producto[0], producto[1]])

        # Crear gráfico
        chart = BarChart()
        chart.title = "Top Productos Menos Vendidos"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(productos) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(productos) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E5")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="reporte_productos_menos_vendidos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template('form_reporte.html', tipo='Productos Menos Vendidos')

@movimientos_bp.route('/reporte-mejores-vendedores', methods=['GET', 'POST'])
def reporte_mejores_vendedores():
    if request.method == 'POST':
        mes = request.form['mes']
        anio = request.form['anio']
        
        vendedores = db.session.execute(text("""
            EXEC sp_ObtenerMejorVendedorPorMesYAño :mes, :anio
        """), {'mes': mes, 'anio': anio}).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Mejores Vendedores"
        ws.append(["Vendedor", "Cantidad Vendida"])

        for vendedor in vendedores:
            ws.append([vendedor[0], vendedor[1]])

        # Crear gráfico
        chart = BarChart()
        chart.title = "Top Mejores Vendedores"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(vendedores) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(vendedores) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E5")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="reporte_mejores_vendedores.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template('form_reporte.html', tipo='Mejores Vendedores')

@movimientos_bp.route('/reporte-movimientos-por-mes', methods=['GET', 'POST'])
def reporte_movimientos_por_mes():
    if request.method == 'POST':
        anio = request.form['anio']
        
        movimientos = db.session.execute(text("""
            EXEC sp_ObtenerMesesConMasMovimientos :anio
        """), {'anio': anio}).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos por Mes"
        ws.append(["Tipo de Movimiento", "Cantidad"])

        for movimiento in movimientos:
            ws.append([movimiento[0], movimiento[1]])

        # Crear gráfico
        chart = BarChart()
        chart.title = "Movimientos por Mes"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(movimientos) + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(movimientos) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "E5")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="reporte_movimientos_por_mes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template('form_reporte.html', tipo='Movimientos por Mes')
